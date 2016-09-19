"""
@author: yosi setiawan
Modified by yanuar harry
signed-off yosi setiawan
signed-off yanuar harry

License GPLv3
"""

import re
import os
import errno


def make_sure_path_exist(path):
    try:
        os.makedirs(path)
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise


def clear_all():
    """Clears all the variables from the workspace of the spyder application."""
    gl = globals().copy()
    for var in gl:
        if var[0] == '_':
            continue
        if 'func' in str(globals()[var]):
            continue
        if 'module' in str(globals()[var]):
            continue

        del globals()[var]
if __name__ == "__main__":
    clear_all()

clear_all()

from datetime import datetime
from datetime import timedelta
import time
import numpy as np
import re
import string
import calendar
import os
from openpyxl import Workbook

date1 = datetime.utcnow()
tahun1 = int(date1.year)
bulan1 = int(date1.month)
tanggal1 = int(date1.day)
if tanggal1 == 1:
    date1 = datetime.utcnow() - timedelta(days=1)
    tahun1 = int(date1.year)
    bulan1 = int(date1.month)

if bulan1 == 1:
    hari_bulan = 31
elif bulan1 == 2:
    if calendar.isleap(tahun1) == True:
        hari_bulan = 29
    else:
        hari_bulan = 28
elif bulan1 == 3:
    hari_bulan = 31
elif bulan1 == 4:
    hari_bulan = 30
elif bulan1 == 5:
    hari_bulan = 31
elif bulan1 == 6:
    hari_bulan = 30
elif bulan1 == 7:
    hari_bulan = 31
elif bulan1 == 8:
    hari_bulan = 31
elif bulan1 == 9:
    hari_bulan = 30
elif bulan1 == 10:
    hari_bulan = 31
elif bulan1 == 11:
    hari_bulan = 30
elif bulan1 == 12:
    hari_bulan = 31

Bx = list(np.tile(np.nan, 1440))
By = list(np.tile(np.nan, 1440))
Bz = list(np.tile(np.nan, 1440))
Bf = list(np.tile(np.nan, 1440))
Bh = list(np.tile(np.nan, 1440))
Bd = list(np.tile(np.nan, 1440))
Bi = list(np.tile(np.nan, 1440))
x_mean = [['AM' for x in range(24)] for x in range(hari_bulan)]
y_mean = [['AM' for x in range(24)] for x in range(hari_bulan)]
z_mean = [['AM' for x in range(24)] for x in range(hari_bulan)]
f_mean = [['AM' for x in range(24)] for x in range(hari_bulan)]
h_mean = [['AM' for x in range(24)] for x in range(hari_bulan)]
d_mean = [['AM' for x in range(24)] for x in range(hari_bulan)]
i_mean = [['AM' for x in range(24)] for x in range(hari_bulan)]
k_i = ['AM'] * hari_bulan * 8
a_i = [-1] * hari_bulan * 8
sk = [''] * hari_bulan
A_i = [-1] * hari_bulan
status = [''] * hari_bulan

wb = Workbook()
namefolder = '/home/tuntsi/Magnet_dev/Excel/' + str(tahun1)
make_sure_path_exist(namefolder)
fileout = namefolder + '/' + string.upper(date1.strftime("%b%Y")) + '.xlsx'

ws1 = wb.active
ws1.merge_cells('A2:Z2')
ws1.merge_cells('B4:B5')
ws1.merge_cells('C4:C5')
ws1.merge_cells('D4:D5')
ws1.merge_cells('E4:E5')
ws1.merge_cells('F4:F5')
ws1.merge_cells('G4:G5')
ws1.merge_cells('H4:H5')
ws1.merge_cells('I4:I5')
ws1.merge_cells('J4:J5')
ws1.merge_cells('K4:K5')
ws1.merge_cells('L4:L5')
ws1.merge_cells('M4:M5')
ws1.merge_cells('N4:N5')
ws1.merge_cells('O4:O5')
ws1.merge_cells('P4:P5')
ws1.merge_cells('Q4:Q5')
ws1.merge_cells('R4:R5')
ws1.merge_cells('S4:S5')
ws1.merge_cells('T4:T5')
ws1.merge_cells('U4:U5')
ws1.merge_cells('V4:V5')
ws1.merge_cells('W4:W5')
ws1.merge_cells('X4:X5')
ws1.merge_cells('Y4:Y5')
ws1.merge_cells('Z4:Z5')
ws1['A2'] = 'Hourly Mean Values of X in nT From Digital Magnet'
ws1['A3'] = 'Tuntungan Magnetic Observatory'
ws1['A4'] = 'Hour'
ws1['A5'] = 'Day'
ws1['Y3'] = date1.strftime("%B %Y")
ws1['B4'] = '1'
ws1['C4'] = '2'
ws1[
    'D4'] = '3'
ws1['E4'] = '4'
ws1['F4'] = '5'
ws1['G4'] = '6'
ws1['H4'] = '7'
ws1[
    'I4'] = '8'
ws1['J4'] = '9'
ws1['K4'] = '10'
ws1['L4'] = '11'
ws1['M4'] = '12'
ws1[
    'N4'] = '13'
ws1['O4'] = '14'
ws1['P4'] = '15'
ws1['Q4'] = '16'
ws1['R4'] = '17'
ws1[
    'S4'] = '18'
ws1['T4'] = '19'
ws1['U4'] = '20'
ws1['V4'] = '21'
ws1['W4'] = '22'
ws1['X4'] = '23'
ws1['Y4'] = '24'
ws1.title = "X"

ws2 = wb.create_sheet()
ws2.merge_cells('A2:Z2')
ws2.merge_cells('B4:B5')
ws2.merge_cells('C4:C5')
ws2.merge_cells('D4:D5')
ws2.merge_cells('E4:E5')
ws2.merge_cells('F4:F5')
ws2.merge_cells('G4:G5')
ws2.merge_cells('H4:H5')
ws2.merge_cells('I4:I5')
ws2.merge_cells('J4:J5')
ws2.merge_cells('K4:K5')
ws2.merge_cells('L4:L5')
ws2.merge_cells('M4:M5')
ws2.merge_cells('N4:N5')
ws2.merge_cells('O4:O5')
ws2.merge_cells('P4:P5')
ws2.merge_cells('Q4:Q5')
ws2.merge_cells('R4:R5')
ws2.merge_cells('S4:S5')
ws2.merge_cells('T4:T5')
ws2.merge_cells('U4:U5')
ws2.merge_cells('V4:V5')
ws2.merge_cells('W4:W5')
ws2.merge_cells('X4:X5')
ws2.merge_cells('Y4:Y5')
ws2.merge_cells('Z4:Z5')
ws2['A2'] = 'Hourly Mean Values of Y in nT From Digital Magnet'
ws2['A3'] = 'Tuntungan Magnetic Observatory'
ws2['A4'] = 'Hour'
ws2['A5'] = 'Day'
ws2['Y3'] = date1.strftime("%B %Y")
ws2['B4'] = '1'
ws2['C4'] = '2'
ws2[
    'D4'] = '3'
ws2['E4'] = '4'
ws2['F4'] = '5'
ws2['G4'] = '6'
ws2['H4'] = '7'
ws2[
    'I4'] = '8'
ws2['J4'] = '9'
ws2['K4'] = '10'
ws2['L4'] = '11'
ws2['M4'] = '12'
ws2[
    'N4'] = '13'
ws2['O4'] = '14'
ws2['P4'] = '15'
ws2['Q4'] = '16'
ws2['R4'] = '17'
ws2[
    'S4'] = '18'
ws2['T4'] = '19'
ws2['U4'] = '20'
ws2['V4'] = '21'
ws2['W4'] = '22'
ws2['X4'] = '23'
ws2['Y4'] = '24'
ws2.title = "Y"

ws3 = wb.create_sheet()
ws3.merge_cells('A2:Z2')
ws3.merge_cells('B4:B5')
ws3.merge_cells('C4:C5')
ws3.merge_cells('D4:D5')
ws3.merge_cells('E4:E5')
ws3.merge_cells('F4:F5')
ws3.merge_cells('G4:G5')
ws3.merge_cells('H4:H5')
ws3.merge_cells('I4:I5')
ws3.merge_cells('J4:J5')
ws3.merge_cells('K4:K5')
ws3.merge_cells('L4:L5')
ws3.merge_cells('M4:M5')
ws3.merge_cells('N4:N5')
ws3.merge_cells('O4:O5')
ws3.merge_cells('P4:P5')
ws3.merge_cells('Q4:Q5')
ws3.merge_cells('R4:R5')
ws3.merge_cells('S4:S5')
ws3.merge_cells('T4:T5')
ws3.merge_cells('U4:U5')
ws3.merge_cells('V4:V5')
ws3.merge_cells('W4:W5')
ws3.merge_cells('X4:X5')
ws3.merge_cells('Y4:Y5')
ws3.merge_cells('Z4:Z5')
ws3['A2'] = 'Hourly Mean Values of Z in nT From Digital Magnet'
ws3['A3'] = 'Tuntungan Magnetic Observatory'
ws3['A4'] = 'Hour'
ws3['A5'] = 'Day'
ws3['Y3'] = date1.strftime("%B %Y")
ws3['B4'] = '1'
ws3['C4'] = '2'
ws3[
    'D4'] = '3'
ws3['E4'] = '4'
ws3['F4'] = '5'
ws3['G4'] = '6'
ws3['H4'] = '7'
ws3[
    'I4'] = '8'
ws3['J4'] = '9'
ws3['K4'] = '10'
ws3['L4'] = '11'
ws3['M4'] = '12'
ws3[
    'N4'] = '13'
ws3['O4'] = '14'
ws3['P4'] = '15'
ws3['Q4'] = '16'
ws3['R4'] = '17'
ws3[
    'S4'] = '18'
ws3['T4'] = '19'
ws3['U4'] = '20'
ws3['V4'] = '21'
ws3['W4'] = '22'
ws3['X4'] = '23'
ws3['Y4'] = '24'
ws3.title = "Z"

ws4 = wb.create_sheet()
ws4.merge_cells('A2:Z2')
ws4.merge_cells('B4:B5')
ws4.merge_cells('C4:C5')
ws4.merge_cells('D4:D5')
ws4.merge_cells('E4:E5')
ws4.merge_cells('F4:F5')
ws4.merge_cells('G4:G5')
ws4.merge_cells('H4:H5')
ws4.merge_cells('I4:I5')
ws4.merge_cells('J4:J5')
ws4.merge_cells('K4:K5')
ws4.merge_cells('L4:L5')
ws4.merge_cells('M4:M5')
ws4.merge_cells('N4:N5')
ws4.merge_cells('O4:O5')
ws4.merge_cells('P4:P5')
ws4.merge_cells('Q4:Q5')
ws4.merge_cells('R4:R5')
ws4.merge_cells('S4:S5')
ws4.merge_cells('T4:T5')
ws4.merge_cells('U4:U5')
ws4.merge_cells('V4:V5')
ws4.merge_cells('W4:W5')
ws4.merge_cells('X4:X5')
ws4.merge_cells('Y4:Y5')
ws4.merge_cells('Z4:Z5')
ws4['A2'] = 'Hourly Mean Values of H in nT From Digital Magnet'
ws4['A3'] = 'Tuntungan Magnetic Observatory'
ws4['A4'] = 'Hour'
ws4['A5'] = 'Day'
ws4['Y3'] = date1.strftime("%B %Y")
ws4['B4'] = '1'
ws4['C4'] = '2'
ws4[
    'D4'] = '3'
ws4['E4'] = '4'
ws4['F4'] = '5'
ws4['G4'] = '6'
ws4['H4'] = '7'
ws4[
    'I4'] = '8'
ws4['J4'] = '9'
ws4['K4'] = '10'
ws4['L4'] = '11'
ws4['M4'] = '12'
ws4[
    'N4'] = '13'
ws4['O4'] = '14'
ws4['P4'] = '15'
ws4['Q4'] = '16'
ws4['R4'] = '17'
ws4[
    'S4'] = '18'
ws4['T4'] = '19'
ws4['U4'] = '20'
ws4['V4'] = '21'
ws4['W4'] = '22'
ws4['X4'] = '23'
ws4['Y4'] = '24'
ws4.title = "H"

ws5 = wb.create_sheet()
ws5.merge_cells('A2:Z2')
ws5.merge_cells('B4:B5')
ws5.merge_cells('C4:C5')
ws5.merge_cells('D4:D5')
ws5.merge_cells('E4:E5')
ws5.merge_cells('F4:F5')
ws5.merge_cells('G4:G5')
ws5.merge_cells('H4:H5')
ws5.merge_cells('I4:I5')
ws5.merge_cells('J4:J5')
ws5.merge_cells('K4:K5')
ws5.merge_cells('L4:L5')
ws5.merge_cells('M4:M5')
ws5.merge_cells('N4:N5')
ws5.merge_cells('O4:O5')
ws5.merge_cells('P4:P5')
ws5.merge_cells('Q4:Q5')
ws5.merge_cells('R4:R5')
ws5.merge_cells('S4:S5')
ws5.merge_cells('T4:T5')
ws5.merge_cells('U4:U5')
ws5.merge_cells('V4:V5')
ws5.merge_cells('W4:W5')
ws5.merge_cells('X4:X5')
ws5.merge_cells('Y4:Y5')
ws5.merge_cells('Z4:Z5')
ws5['A2'] = 'Hourly Mean Values of F in nT From Digital Magnet'
ws5['A3'] = 'Tuntungan Magnetic Observatory'
ws5['A4'] = 'Hour'
ws5['A5'] = 'Day'
ws5['Y3'] = date1.strftime("%B %Y")
ws5['B4'] = '1'
ws5['C4'] = '2'
ws5[
    'D4'] = '3'
ws5['E4'] = '4'
ws5['F4'] = '5'
ws5['G4'] = '6'
ws5['H4'] = '7'
ws5[
    'I4'] = '8'
ws5['J4'] = '9'
ws5['K4'] = '10'
ws5['L4'] = '11'
ws5['M4'] = '12'
ws5[
    'N4'] = '13'
ws5['O4'] = '14'
ws5['P4'] = '15'
ws5['Q4'] = '16'
ws5['R4'] = '17'
ws5[
    'S4'] = '18'
ws5['T4'] = '19'
ws5['U4'] = '20'
ws5['V4'] = '21'
ws5['W4'] = '22'
ws5['X4'] = '23'
ws5['Y4'] = '24'
ws5.title = "F"

ws6 = wb.create_sheet()
ws6.merge_cells('A2:Z2')
ws6.merge_cells('B4:B5')
ws6.merge_cells('C4:C5')
ws6.merge_cells('D4:D5')
ws6.merge_cells('E4:E5')
ws6.merge_cells('F4:F5')
ws6.merge_cells('G4:G5')
ws6.merge_cells('H4:H5')
ws6.merge_cells('I4:I5')
ws6.merge_cells('J4:J5')
ws6.merge_cells('K4:K5')
ws6.merge_cells('L4:L5')
ws6.merge_cells('M4:M5')
ws6.merge_cells('N4:N5')
ws6.merge_cells('O4:O5')
ws6.merge_cells('P4:P5')
ws6.merge_cells('Q4:Q5')
ws6.merge_cells('R4:R5')
ws6.merge_cells('S4:S5')
ws6.merge_cells('T4:T5')
ws6.merge_cells('U4:U5')
ws6.merge_cells('V4:V5')
ws6.merge_cells('W4:W5')
ws6.merge_cells('X4:X5')
ws6.merge_cells('Y4:Y5')
ws6.merge_cells('Z4:Z5')
ws6['A2'] = 'Hourly Mean Values of D in arcMin From Digital Magnet'
ws6['A3'] = 'Tuntungan Magnetic Observatory'
ws6['A4'] = 'Hour'
ws6['A5'] = 'Day'
ws6['Y3'] = date1.strftime("%B %Y")
ws6['B4'] = '1'
ws6['C4'] = '2'
ws6[
    'D4'] = '3'
ws6['E4'] = '4'
ws6['F4'] = '5'
ws6['G4'] = '6'
ws6['H4'] = '7'
ws6[
    'I4'] = '8'
ws6['J4'] = '9'
ws6['K4'] = '10'
ws6['L4'] = '11'
ws6['M4'] = '12'
ws6[
    'N4'] = '13'
ws6['O4'] = '14'
ws6['P4'] = '15'
ws6['Q4'] = '16'
ws6['R4'] = '17'
ws6[
    'S4'] = '18'
ws6['T4'] = '19'
ws6['U4'] = '20'
ws6['V4'] = '21'
ws6['W4'] = '22'
ws6['X4'] = '23'
ws6['Y4'] = '24'
ws6.title = "D"

ws7 = wb.create_sheet()
ws7.merge_cells('A2:Z2')
ws7.merge_cells('B4:B5')
ws7.merge_cells('C4:C5')
ws7.merge_cells('D4:D5')
ws7.merge_cells('E4:E5')
ws7.merge_cells('F4:F5')
ws7.merge_cells('G4:G5')
ws7.merge_cells('H4:H5')
ws7.merge_cells('I4:I5')
ws7.merge_cells('J4:J5')
ws7.merge_cells('K4:K5')
ws7.merge_cells('L4:L5')
ws7.merge_cells('M4:M5')
ws7.merge_cells('N4:N5')
ws7.merge_cells('O4:O5')
ws7.merge_cells('P4:P5')
ws7.merge_cells('Q4:Q5')
ws7.merge_cells('R4:R5')
ws7.merge_cells('S4:S5')
ws7.merge_cells('T4:T5')
ws7.merge_cells('U4:U5')
ws7.merge_cells('V4:V5')
ws7.merge_cells('W4:W5')
ws7.merge_cells('X4:X5')
ws7.merge_cells('Y4:Y5')
ws7.merge_cells('Z4:Z5')
ws7['A2'] = 'Hourly Mean Values of I in Degree From Digital Magnet'
ws7['A3'] = 'Tuntungan Magnetic Observatory'
ws7['A4'] = 'Hour'
ws7['A5'] = 'Day'
ws7['Y3'] = date1.strftime("%B %Y")
ws7['B4'] = '1'
ws7['C4'] = '2'
ws7[
    'D4'] = '3'
ws7['E4'] = '4'
ws7['F4'] = '5'
ws7['G4'] = '6'
ws7['H4'] = '7'
ws7[
    'I4'] = '8'
ws7['J4'] = '9'
ws7['K4'] = '10'
ws7['L4'] = '11'
ws7['M4'] = '12'
ws7[
    'N4'] = '13'
ws7['O4'] = '14'
ws7['P4'] = '15'
ws7['Q4'] = '16'
ws7['R4'] = '17'
ws7[
    'S4'] = '18'
ws7['T4'] = '19'
ws7['U4'] = '20'
ws7['V4'] = '21'
ws7['W4'] = '22'
ws7['X4'] = '23'
ws7['Y4'] = '24'
ws7.title = "I"

ws8 = wb.create_sheet()
ws8.merge_cells('A1:K1')
ws8.merge_cells('A8:K8')
ws8.merge_cells('A9:K9')
ws8.merge_cells('A10:A11')
ws8.merge_cells(
    'B10:B11')
ws8.merge_cells('C10:C11')
ws8.merge_cells('D10:D11')
ws8.merge_cells(
    'E10:E11')
ws8.merge_cells('F10:F11')
ws8.merge_cells('G10:G11')
ws8.merge_cells(
    'H10:H11')
ws8.merge_cells('I10:I11')
ws8.merge_cells('J10:J11')
ws8.merge_cells('K10:K11')
ws8['A1'] = 'M A G N E T I C  A C T I V I T Y'
ws8['A4'] = 'Observatory'
ws8['C4'] = ': Tuntungan  - %s  %s' % (
    string.upper(date1.strftime("%B")), tahun1)
ws8['A5'] = 'Geog. Latitude'
ws8[
    'C5'] = ': 03 30 01.4 N'
ws8['E5'] = 'Geom. Latitude'
ws8['I5'] = 'Type of instr  : LEMI-018'
ws8['A6'] = 'Geog. Long.'
ws8[
    'C6'] = ': 98 33 51.6 E'
ws8['E6'] = 'Geom. Longitute'
ws8['A8'] = 'K - Indices for three hours interval'
ws8['A9'] = 'K - 9 = 300 gammas'
ws8['A10'] = 'DATE'
ws8['B10'] = '00-03'
ws8[
    'C10'] = '03-06'
ws8['D10'] = '06-09'
ws8['E10'] = '09-12'
ws8['F10'] = '12-15'
ws8[
    'G10'] = '15-18'
ws8['H10'] = '18-21'
ws8['I10'] = '21-24'
ws8['J10'] = 'SUM'
ws8['K10'] = 'CHARACTER'
ws8.title = "K"

for m in range(0, hari_bulan):
    filename = 'TUN%4.0f%02.0f%02.0fvmin.min' % (tahun1, bulan1, m + 1)
    namefolderiaga = '/home/tuntsi/Magnet_dev/IAGA/' + \
        str(tahun1) + '/' + str(bulan1).zfill(2)

    try:
        f_iaga = open(namefolderiaga + '/' + filename)

        content = f_iaga.readlines()[13:]
        for i in range(0, 1440):
            data = re.split('\s+', content[i])
            if data[3] != '99999.00':
                Bx[i] = float(data[3])
            else:
                Bx[i] = np.nan
            if data[4] != '99999.00':
                By[i] = float(data[4])
            else:
                By[i] = np.nan
            if data[5] != '99999.00':
                Bz[i] = float(data[5])
            else:
                Bz[i] = np.nan
            Bf[i] = np.sqrt((Bx[i] ** 2) + (By[i] ** 2) + (Bz[i] ** 2))
            Bh[i] = np.sqrt((Bx[i] ** 2) + (By[i] ** 2))
            Bd[i] = np.rad2deg(np.arctan(By[i] / Bx[i])) * 60
            Bi[i] = np.rad2deg(np.arctan(Bz[i] / Bh[i]))
    except IOError as err:
        continue

    for k in range(0, 24):
        if np.count_nonzero(np.isnan(Bx[(0 + (60 * k)):((60 + (60 * k)))])) <= 6:
            x_mean[m][k] = '%7.1f' % np.nanmean(
                Bx[(0 + (60 * k)):((60 + (60 * k)))])
        else:
            x_mean[m][k] = 'AM'
        if np.count_nonzero(np.isnan(By[(0 + (60 * k)):((60 + (60 * k)))])) <= 6:
            y_mean[m][k] = '%6.1f' % np.nanmean(
                By[(0 + (60 * k)):((60 + (60 * k)))])
        else:
            y_mean[m][k] = 'AM'
        if np.count_nonzero(np.isnan(Bz[(0 + (60 * k)):((60 + (60 * k)))])) <= 6:
            z_mean[m][k] = '%7.1f' % np.nanmean(
                Bz[(0 + (60 * k)):((60 + (60 * k)))])
        else:
            z_mean[m][k] = 'AM'
        if np.count_nonzero(np.isnan(Bf[(0 + (60 * k)):((60 + (60 * k)))])) <= 6:
            f_mean[m][k] = '%7.1f' % np.nanmean(
                Bf[(0 + (60 * k)):((60 + (60 * k)))])
        else:
            f_mean[m][k] = 'AM'
        if np.count_nonzero(np.isnan(Bh[(0 + (60 * k)):((60 + (60 * k)))])) <= 6:
            h_mean[m][k] = '%7.1f' % np.nanmean(
                Bh[(0 + (60 * k)):((60 + (60 * k)))])
        else:
            h_mean[m][k] = 'AM'
        if np.count_nonzero(np.isnan(Bd[(0 + (60 * k)):((60 + (60 * k)))])) <= 6:
            d_mean[m][k] = '%6.1f' % np.nanmean(
                Bd[(0 + (60 * k)):((60 + (60 * k)))])
        else:
            d_mean[m][k] = 'AM'
        if np.count_nonzero(np.isnan(Bi[(0 + (60 * k)):((60 + (60 * k)))])) <= 6:
            i_mean[m][k] = '%6.1f' % np.nanmean(
                Bi[(0 + (60 * k)):((60 + (60 * k)))])
        else:
            i_mean[m][k] = 'AM'

namafile4 = '/home/tuntsi/Magnet_dev/k_bulan.sh'
f_k = open(namafile4, 'w')
f_k.write('kasm TUN:01%s:%2.0f 300 xy /home/tuntsi/Magnet_dev/data_bulan /home/tuntsi/Magnet_dev/IMFVALL/' %
          ((string.upper(date1.strftime("%b%Y"))), hari_bulan))
f_k.close()
os.system('sh /home/tuntsi/Magnet_dev/k_bulan.sh')

with open('/home/tuntsi/Magnet_dev/data_bulan.dka') as f:
    content = f.readlines()
    index_k = [-1] * len(content)
    for j in range(6, len(content)):
        data_k = re.split(' ', content[j])
        for i in range(0, len(data_k) - 14):
            data_k.remove('')
        index_kk = time.strptime(data_k[0], '%d-%b-%y').tm_mday - 1
        sk[index_kk] = int(data_k[10])
        if sk[index_kk] == -1:
            sk[index_kk] = ''

        for m in range(0, 8):
            k_i[m + ((index_kk) * 8)] = int(data_k[m + 2])
            if (k_i[m + ((index_kk) * 8)] == 0):
                a_i[m + ((index_kk) * 8)] = 0
            elif (k_i[m + ((index_kk) * 8)] == 1):
                a_i[m + ((index_kk) * 8)] = 3
            elif (k_i[m + ((index_kk) * 8)] == 2):
                a_i[m + ((index_kk) * 8)] = 6
            elif (k_i[m + ((index_kk) * 8)] == 3):
                a_i[m + ((index_kk) * 8)] = 12
            elif (k_i[m + ((index_kk) * 8)] == 4):
                a_i[m + ((index_kk) * 8)] = 24
            elif (k_i[m + ((index_kk) * 8)] == 5):
                a_i[m + ((index_kk) * 8)] = 40
            elif (k_i[m + ((index_kk) * 8)] == 6):
                a_i[m + ((index_kk) * 8)] = 70
            elif (k_i[m + ((index_kk) * 8)] == 7):
                a_i[m + ((index_kk) * 8)] = 120
            elif (k_i[m + ((index_kk) * 8)] == 8):
                a_i[m + ((index_kk) * 8)] = 200
            elif (k_i[m + ((index_kk) * 8)] == 9):
                a_i[m + ((index_kk) * 8)] = 300
            elif (k_i[m + ((index_kk) * 8)] == -1):
                a_i[m + ((index_kk) * 8)] = -1
                k_i[m + ((index_kk) * 8)] = 'AM'


for n in range(0, hari_bulan):
    if sk[n] == '':
        status[n] = ''
    else:
        A_i[n] = sum(a_i[0 + (n * 8):8 + (n * 8)]) / 8
        if A_i < 0:
            status[n] = ''
        elif (A_i[n] >= 0 and A_i[n] <= 30):
            status[n] = 'Hari Tenang'
        elif (A_i[n] > 30 and A_i[n] <= 50):
            status[n] = 'Badai Lemah'
        elif (A_i[n] > 50 and A_i[n] <= 100):
            status[n] = 'Badai Menengah'
        elif (A_i[n] > 100):
            status[n] = 'Badai Kuat'


for rows in range(1, hari_bulan + 1):
    ws1['A' + str(rows + 5)] = rows
    for columns in range(1, 25):
        ws1.cell(column=columns + 1, row=rows + 5,
                 value=('%s' % x_mean[rows - 1][columns - 1]))
    ws2['A' + str(rows + 5)] = rows
    for columns in range(1, 25):
        ws2.cell(column=columns + 1, row=rows + 5,
                 value=('%s' % y_mean[rows - 1][columns - 1]))
    ws3['A' + str(rows + 5)] = rows
    for columns in range(1, 25):
        ws3.cell(column=columns + 1, row=rows + 5,
                 value=('%s' % z_mean[rows - 1][columns - 1]))
    ws4['A' + str(rows + 5)] = rows
    for columns in range(1, 25):
        ws4.cell(column=columns + 1, row=rows + 5,
                 value=('%s' % h_mean[rows - 1][columns - 1]))
    ws5['A' + str(rows + 5)] = rows
    for columns in range(1, 25):
        ws5.cell(column=columns + 1, row=rows + 5,
                 value=('%s' % f_mean[rows - 1][columns - 1]))
    ws6['A' + str(rows + 5)] = rows
    for columns in range(1, 25):
        ws6.cell(column=columns + 1, row=rows + 5,
                 value=('%s' % d_mean[rows - 1][columns - 1]))
    ws7['A' + str(rows + 5)] = rows
    for columns in range(1, 25):
        ws7.cell(column=columns + 1, row=rows + 5,
                 value=('%s' % i_mean[rows - 1][columns - 1]))
    ws8['A' + str(rows + 11)] = rows
    for columns in range(1, 9):
        ws8.cell(column=columns + 1, row=rows + 11,
                 value=('%s' % k_i[(8 * (rows - 1)) - 1 + columns]))
    ws8.cell(column=10, row=rows + 11, value=('%s' % sk[rows - 1]))
    ws8.cell(column=11, row=rows + 11, value=('%s' % status[rows - 1]))

wb.save(filename=fileout)

print '%s has been created' % (fileout)
